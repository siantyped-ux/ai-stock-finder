import sys

import pytest
from openpyxl import load_workbook

import exit_rules as er
import perf_report as pr
import trade_sim as ts


def _trade(**kw):
    """기본은 AAA 를 08-03 @$100 에 사서 08-05 @$110 에 판 트레이드."""
    base = dict(
        ticker="AAA", market="US", source="live",
        entry_date="2026-08-03", entry_price=100.0, r_unit=6.0,
        exit_date="2026-08-05", exit_price=110.0, mark_price=110.0,
        exit_reason="TRAIL", bars_held=2, is_open=False,
        gross_r=1.67, cost_r=0.05, net_r=1.62,
        initial_stop=94.0, high_since_entry=110.0, stop=94.0,
        target_price=None, qty=10,
    )
    base.update(kw)
    return ts.Trade(**base)


# 금액은 전부 달러다. 미국 종목만 보므로 원화 환산은 손익에 아무것도
# 더해 주지 않으면서 USDKRW 조회 실패라는 실패 경로만 만들었다.

# 수량 규칙은 sizing.py 로 옮겨갔다 - 리스크로 역산하고, 못 사면 건너뛴다.
# 여기서는 시뮬레이터가 준 값을 그대로 쓰는지만 본다.

def test_a_trade_converts_at_face_value():
    # 원금 100x10 = 1,000 / 회수 110x10 = 1,100
    # 매수비용 0.15x10 = 1.5 / 매도비용 0.165x10 = 1.65
    row = pr.to_row(_trade())

    assert row["qty"] == 10
    assert row["gross_usd"] == pytest.approx(100.0)
    assert row["gross_pct"] == pytest.approx(10.0)
    assert row["net_usd"] == pytest.approx(96.85)
    assert row["net_pct"] == pytest.approx(9.685)


def test_loss_stays_negative_and_costs_make_it_worse():
    row = pr.to_row(_trade(exit_price=90.0, mark_price=90.0))

    assert row["gross_usd"] < 0
    assert row["net_usd"] < row["gross_usd"]


def test_cost_agrees_with_cost_r():
    # 1주면 금액 비용은 cost_r x r_unit 과 같아야 한다. 요율 분기가 두 곳에
    # 복제되면 이 등식이 깨진다.
    t = _trade(entry_price=500.0, exit_price=550.0, mark_price=550.0,
               r_unit=30.0, qty=1)
    row = pr.to_row(t)

    assert row["qty"] == 1
    expected = ts.cost_r(500.0, 550.0, 30.0, "US", ts.Costs()) * 30.0
    assert row["gross_usd"] - row["net_usd"] == pytest.approx(expected)


def test_open_position_uses_the_mark_price_and_still_pays_the_sell_side():
    t = _trade(is_open=True, exit_date=None, exit_price=None, mark_price=105.0)
    row = pr.to_row(t)

    assert row["exit_price"] == 105.0
    # 매도비용을 빼지 않으면 net == gross 가 된다
    assert row["net_usd"] < row["gross_usd"]


def _write(path, built):
    """주식 트랙만 채워 리포트를 쓴다. 단일 트랙 동작을 보는 테스트가 쓴다."""
    pr.write_xlsx(path, {"stocks": built, "etf": None})


def _result(trades, **kw):
    base = dict(
        trades=trades, dates=["2026-08-03", "2026-08-05"],
        live_rows=10, backfill_rows=90, failed=[],
        newest_bar="2026-08-05", cash=None, capital=None,
    )
    base.update(kw)
    return base


def test_open_positions_never_land_in_the_closed_sheet():
    built = pr.build_rows(_result([
        _trade(),
        _trade(ticker="BBB", is_open=True, exit_date=None,
               exit_price=None, mark_price=105.0),
    ]))

    assert [r["ticker"] for r in built["closed"]] == ["AAA"]
    assert [r["ticker"] for r in built["open"]] == ["BBB"]


def test_open_position_is_marked_to_the_newest_bar_date():
    built = pr.build_rows(_result([
        _trade(is_open=True, exit_date=None, exit_price=None, mark_price=105.0),
    ]))

    assert built["open"][0]["exit_date"] == "2026-08-05"


def test_win_rate_ignores_open_positions():
    # 닫힌 2건 중 1승. 미결은 큰 이익이지만 승률에 들어가면 안 된다 -
    # "아직 손절되지 않았을 뿐" 인 포지션이다.
    built = pr.build_rows(_result([
        _trade(ticker="WIN"),
        _trade(ticker="LOSS", exit_price=90.0, mark_price=90.0),
        _trade(ticker="OPEN", is_open=True, exit_date=None,
               exit_price=None, mark_price=200.0),
    ]))
    s = built["summary"]

    assert s["closed_n"] == 2
    assert s["win_rate"] == pytest.approx(50.0)
    assert s["open_n"] == 1


def test_closed_rows_sort_by_exit_date_then_ticker():
    built = pr.build_rows(_result([
        _trade(ticker="ZZZ", exit_date="2026-08-05"),
        _trade(ticker="AAA", exit_date="2026-08-05"),
        _trade(ticker="MMM", exit_date="2026-08-03"),
    ]))

    assert [r["ticker"] for r in built["closed"]] == ["MMM", "AAA", "ZZZ"]


def test_summary_survives_zero_closed_trades():
    s = pr.build_rows(_result([]))["summary"]

    assert s["closed_n"] == 0
    assert s["win_rate"] is None
    assert s["avg_net_pct"] is None


def test_closed_sheet_leads_with_the_requested_columns(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()])))

    header = [c.value for c in load_workbook(path)["주식 청산완료"][1]]
    # 환율 열은 사라졌다(미국 종목만 보므로 금액이 전부 달러). 시장 열이
    # 티커 옆에 붙었다.
    assert header == ["상품티커", "시장", "진입일자", "진입가격",
                      "청산일자", "청산가격", "총수익($)", "총수익(%)",
                      "순수익($)", "순수익(%)", "수량", "청산사유"]


def test_negative_money_renders_with_a_minus_sign(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(
        _result([_trade(exit_price=90.0, mark_price=90.0)])))

    cell = _cell(path, "주식 청산완료", "총수익($)")

    assert cell.value < 0
    assert cell.number_format == "#,##0.00;-#,##0.00"
    # 회계 서식의 괄호 표기여서는 안 된다
    assert "(" not in cell.number_format


def test_percent_cells_store_the_readable_number_not_a_fraction(tmp_path):
    # 값이 0.1423 이면 셀을 직접 읽는 쪽이 100배 틀린다.
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()])))

    cell = _cell(path, "주식 청산완료", "총수익(%)")

    assert cell.value > 1.0
    assert cell.number_format == '0.00"%";-0.00"%"'


def test_closed_sheet_keeps_its_header_when_there_are_no_trades(tmp_path):
    # 시트가 없으면 파일이 깨진 것인지 트레이드가 없는 것인지 구분되지 않는다.
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([])))

    ws = load_workbook(path)["주식 청산완료"]

    assert ws.max_row == 1
    assert ws["A1"].value == "상품티커"


def test_open_sheet_labels_the_valuation_columns(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([
        _trade(is_open=True, exit_date=None, exit_price=None, mark_price=105.0),
    ])))

    header = [c.value for c in load_workbook(path)["주식 미결포지션"][1]]

    # 평가 시점 열이 청산 시트의 청산일자·청산가격 자리를 대신한다.
    assert "평가기준일" in header and "현재가" in header
    assert header.index("평가기준일") < header.index("현재가")
    assert "보유봉수" in header


def test_summary_leads_with_the_contamination_warning(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()])))

    ws = load_workbook(path)["요약"]

    assert ws["A1"].value == "!! 경고"
    assert "파이프라인 검증용" in ws["B1"].value


def _cell(path, sheet, header, row=2):
    """헤더 이름으로 셀을 찾는다. 열 위치가 바뀌어도 테스트가 안 깨진다."""
    ws = load_workbook(path)[sheet]
    names = [c.value for c in ws[1]]
    return ws.cell(row=row, column=names.index(header) + 1)


def _held(**kw):
    """미결 포지션 기본형. 진입 100 · 1R=6 · 손절 94 · 최고 102 · 평가 101.5."""
    base = dict(is_open=True, exit_date=None, exit_price=None,
                mark_price=101.5, exit_reason=None,
                initial_stop=94.0, high_since_entry=102.0, stop=94.0)
    base.update(kw)
    return _trade(**base)


def test_open_sheet_appends_the_stop_columns(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_held()])))

    header = [c.value for c in load_workbook(path)["주식 미결포지션"][1]]

    tail = header[header.index("손절가"):]
    assert tail[:4] == ["손절가", "손절까지(%)", "트레일 발동가", "트레일"]


def test_stop_distance_is_negative_because_it_is_a_drop(tmp_path):
    # (101.5 - 94) / 101.5 = 7.39% 아래. 부호가 없으면 상승 여력처럼 읽힌다.
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_held()])))

    assert _cell(path, "주식 미결포지션", "손절가").value == pytest.approx(94.0)
    room = _cell(path, "주식 미결포지션", "손절까지(%)")
    assert room.value == pytest.approx(-7.3892, abs=1e-4)
    assert room.number_format == '0.00"%";-0.00"%"'


def test_trail_trigger_is_entry_plus_one_r(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_held()])))

    assert _cell(path, "주식 미결포지션", "트레일 발동가").value == pytest.approx(106.0)
    assert _cell(path, "주식 미결포지션", "트레일").value == "off"   # 최고 102 < 106


def test_trail_shows_on_once_the_high_reaches_the_trigger(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(
        _result([_held(high_since_entry=106.0, stop=100.0)])))

    assert _cell(path, "주식 미결포지션", "트레일").value == "ON"
    # 본전으로 올라온 손절선
    assert _cell(path, "주식 미결포지션", "손절가").value == pytest.approx(100.0)


def test_closed_sheet_is_untouched_by_the_stop_columns(tmp_path):
    # 청산된 트레이드에는 "지금 어디서 잘리는가" 가 없다.
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()])))

    header = [c.value for c in load_workbook(path)["주식 청산완료"][1]]

    assert len(header) == 12
    assert "손절가" not in header


def test_open_sheet_appends_the_target_columns(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(
        _result([_held(target_price=109.0)])))

    header = [c.value for c in load_workbook(path)["주식 미결포지션"][1]]

    assert header[-4:] == ["목표(%)", "목표가", "달성률(%)", "위험보상"]


def test_target_progress_measures_the_way_to_the_target():
    # 진입 100 · 목표 109 · 평가 101.5 → 목표폭 9 중 1.5 만큼 왔다.
    # 위험보상 = 목표폭 9 / 1R 6 = 1.5
    row = pr.build_rows(_result([_held(target_price=109.0)]))["open"][0]

    assert row["target_pct"] == pytest.approx(9.0)
    assert row["target_price"] == pytest.approx(109.0)
    assert row["target_progress_pct"] == pytest.approx(16.6667, abs=1e-4)
    assert row["reward_risk"] == pytest.approx(1.5)


def test_target_progress_is_negative_below_the_entry():
    # 진입가 아래면 목표에서 멀어진 것이다. 부호가 없으면 진행처럼 읽힌다.
    row = pr.build_rows(
        _result([_held(mark_price=97.0, target_price=109.0)]))["open"][0]

    assert row["target_progress_pct"] == pytest.approx(-33.3333, abs=1e-4)


def test_target_columns_are_blank_without_a_target():
    row = pr.build_rows(_result([_held()]))["open"][0]

    assert row["target_pct"] is None
    assert row["target_price"] is None
    assert row["target_progress_pct"] is None
    assert row["reward_risk"] is None


def test_closed_sheet_is_untouched_by_the_target_columns(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(
        _result([_trade(target_price=109.0)])))

    header = [c.value for c in load_workbook(path)["주식 청산완료"][1]]

    assert "목표가" not in header
    assert header[-1] == "청산사유"


def _summary_rows(path):
    return {r[0].value: r[1].value
            for r in load_workbook(path)["요약"].iter_rows(min_col=1,
                                                          max_col=2)}


def test_summary_says_the_target_exit_is_off_by_default(tmp_path):
    # 컬럼은 항상 보이는데 규칙은 꺼져 있다. 어느 쪽인지 적어 두지 않으면
    # 목표가가 익절 예고처럼 읽힌다.
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()])))

    assert _summary_rows(path)["목표가 익절"] == "사용 안 함 (--use-target 으로 켬)"


def test_summary_says_the_target_exit_is_on_when_enabled(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()]),
                                      params=er.Params(use_target=True)))

    assert _summary_rows(path)["목표가 익절"] == "사용함 (목표가 도달 시 청산)"


SUMMARY = {
    "generated": "2026-08-20 10:00 KST",
    "archive_from": "2026-08-01", "archive_to": "2026-08-19",
    "live_rows": 300, "backfill_rows": 800, "backfill_pct": 72.7,
    "mark_date": "2026-08-19", "failed": [],
    "closed_n": 12, "win_rate": 58.333,
    "gross_usd": 1500.0, "net_usd": 1180.0, "avg_net_pct": 2.1,
    "open_n": 8, "open_net_usd": 2240.0, "capital": 1000,
}


def _only(summary):
    """주식 트랙만 채운 리포트 구조. summary_text 가 이 모양을 받는다."""
    return {"stocks": {"summary": summary, "closed": [], "open": []},
            "etf": None}


def test_summary_text_totals_realised_and_unrealised():
    text = pr.summary_text(_only(SUMMARY))

    # 총 손익 = 실현 + 미실현
    assert "$+3,420.00" in text
    assert "$+1,180.00" in text
    assert "$+2,240.00" in text


def test_summary_text_carries_the_closed_stats():
    text = pr.summary_text(_only(SUMMARY))

    assert "청산 12건" in text
    assert "58.3%" in text
    assert "+2.10%" in text
    assert "보유 8종목" in text


def test_summary_text_survives_zero_closed_trades():
    s = dict(SUMMARY, closed_n=0, win_rate=None, avg_net_pct=None,
             net_usd=0)

    # None 을 포맷하면 터진다. 청산 표본이 없는 지금이 바로 그 상태다.
    text = pr.summary_text(_only(s))
    assert "청산 0건" in text
    assert "None" not in text


def test_summary_text_lists_the_failed_tickers():
    assert "없음" in pr.summary_text(_only(SUMMARY))
    assert "AAA, BBB" in pr.summary_text(_only(dict(SUMMARY, failed=["AAA", "BBB"])))


def test_summary_text_leads_with_the_warning():
    text = pr.summary_text(_only(SUMMARY))

    # 메일은 첨부를 안 열고 본문만 훑게 만든다. 경고가 첨부 안에만
    # 있으면 없는 것과 같다.
    assert text.startswith("!!")
    assert "파이프라인 검증용" in text
    assert "73% 가 backfill" in text
    assert "60거래일" in text


def test_summary_text_renders_losses_with_a_minus_sign():
    # 성과리포트의 핵심은 손실이다. +,.0f 서식의 부호 처리는 양수만
    # 테스트해서는 검증되지 않는다.
    s = dict(SUMMARY, net_usd=-1180.0, open_net_usd=-2240.0)
    text = pr.summary_text(_only(s))

    assert "$-3,420.00" in text
    assert "$-1,180.00" in text
    assert "$-2,240.00" in text


def _stub_run(*_a, **_kw):
    """backtest.run() 최소 결과. 트레이드 0건이라 청산 0건 경로도 탄다."""
    return _result([])


def test_mail_flag_sends_the_same_body_that_gets_printed(monkeypatch, tmp_path,
                                                          capsys):
    sent = []
    creds = {"to": "a@b.c", "user": "a@b.c", "password": "x"}
    monkeypatch.setattr(pr.backtest, "run", _stub_run)
    monkeypatch.setattr(pr.mailer, "send",
                        lambda *a, **kw: sent.append((a, kw)))
    # 실 환경변수가 없어도 되도록 자격증명 조회 자체를 스텁한다.
    monkeypatch.setattr(pr.mailer, "creds_from_env", lambda: creds)
    monkeypatch.setattr(sys, "argv",
                        ["perf_report.py", "--mail", "--out-dir", str(tmp_path)])

    pr.main()

    printed = capsys.readouterr().out
    assert len(sent) == 1
    (subject, body, attachments), kw = sent[0]
    # 리포트가 한 통이므로 제목에 트랙을 넣지 않는다.
    assert subject.startswith("[성과리포트]")

    written = list(tmp_path.glob("*.xlsx"))
    assert len(written) == 1
    assert written[0].name.startswith("perf_2026")

    # 메일 본문은 "콘솔에 찍힌 본문 + 첨부 안내" 여야 한다. 둘이 갈라지면
    # summary_text() 를 공유하는 이유가 사라진다 - 부분 문자열 두 개가
    # 우연히 들어 있는 걸로는 이 관계를 못 잡는다.
    suffix = f"\n\n상세는 첨부된 {written[0].name} 참고"
    assert body.endswith(suffix)
    printed_body = body[:-len(suffix)]
    assert printed_body in printed

    # creds_from_env() 가 실제로 send() 까지 전달되는지 확인한다.
    # 스텁이 아무 kwargs 나 받아주므로, 여기서 안 짚으면 **creds_from_env()
    # 가 통째로 빠져도 테스트가 통과한다.
    assert kw == creds

    assert attachments == [written[0]]


def test_without_mail_flag_nothing_gets_sent(monkeypatch, tmp_path):
    sent = []
    monkeypatch.setattr(pr.backtest, "run", _stub_run)
    monkeypatch.setattr(pr.mailer, "send",
                        lambda *a, **kw: sent.append((a, kw)))
    monkeypatch.setattr(sys, "argv",
                        ["perf_report.py", "--out-dir", str(tmp_path)])

    pr.main()

    assert sent == []


# ─── 트랙 분할 리포트 ─────────────────────────────────────────
# 주식과 ETF 는 점수 척도가 달라 성과도 한 표에 섞으면 승률·평균이 무엇의
# 값인지 알 수 없게 된다. 시트를 나누고 요약에서만 합계를 낸다.

def _both(stock_trades, etf_trades):
    return {"stocks": pr.build_rows(_result(stock_trades)),
            "etf": pr.build_rows(_result(etf_trades))}


def test_report_has_a_sheet_per_track_plus_one_summary(tmp_path):
    path = tmp_path / "r.xlsx"
    pr.write_xlsx(path, _both([_trade(ticker="AAA")], [_trade(ticker="SPY")]))

    assert load_workbook(path).sheetnames == [
        "주식 청산완료", "ETF 청산완료",
        "주식 미결포지션", "ETF 미결포지션", "요약"]


def test_each_track_sheet_holds_only_its_own_rows(tmp_path):
    path = tmp_path / "r.xlsx"
    pr.write_xlsx(path, _both([_trade(ticker="AAA")], [_trade(ticker="SPY")]))

    wb = load_workbook(path)
    assert [r[0].value for r in wb["주식 청산완료"].iter_rows(min_row=2)] == ["AAA"]
    assert [r[0].value for r in wb["ETF 청산완료"].iter_rows(min_row=2)] == ["SPY"]


def test_an_empty_track_still_gets_its_sheet(tmp_path):
    # 시트가 없으면 파일이 깨진 것인지 트레이드가 없는 것인지 구분되지 않는다.
    path = tmp_path / "r.xlsx"
    pr.write_xlsx(path, _both([_trade()], []))

    ws = load_workbook(path)["ETF 청산완료"]
    assert ws.max_row == 1
    assert ws["A1"].value == "상품티커"


def test_summary_reports_each_track_and_the_total(tmp_path):
    path = tmp_path / "r.xlsx"
    pr.write_xlsx(path, _both(
        [_trade(ticker="AAA")],
        [_trade(ticker="SPY", exit_price=90.0, mark_price=90.0)]))

    labels = [r[0].value for r in load_workbook(path)["요약"].iter_rows()]
    assert "[주식]" in labels
    assert "[ETF]" in labels
    assert "[합계]" in labels


def test_summary_text_totals_both_tracks():
    built = _both([_trade(ticker="AAA")], [_trade(ticker="SPY")])
    text = pr.summary_text(built)

    both = (built["stocks"]["summary"]["net_usd"]
            + built["etf"]["summary"]["net_usd"])
    assert f"${both:+,.2f}" in text
    assert "주식" in text and "ETF" in text


# ─── 통화 혼합 방지 ───────────────────────────────────────────
# 아카이브 앞부분(2026-07-31~08-21)에는 원화로 호가되는 한국 종목이 남아
# 있다. 환율을 걷어낸 뒤로 그 가격을 그대로 달러로 셈하면 GS 의 20,900원
# 손익이 $20,900 이 된다 - 실제로 한 번 그렇게 나왔다.

def test_a_non_usd_trade_is_refused():
    with pytest.raises(ValueError, match="US"):
        pr.to_row(_trade(market="KR", entry_price=95100.0,
                         exit_price=116000.0, mark_price=116000.0))


def test_the_report_asks_the_backtest_for_us_rows_only(monkeypatch):
    seen = {}

    def spy(pattern, params=None, **kw):
        seen[pattern] = kw
        return _result([])

    monkeypatch.setattr(pr.backtest, "run", spy)
    monkeypatch.setattr(sys, "argv", ["perf_report.py", "--out-dir", "."])
    try:
        pr.main()
    except SystemExit:
        pass

    assert seen, "백테스트를 부르지 않았다"
    assert all(kw.get("us_only") for kw in seen.values())
    # 집계는 정해진 날짜부터다. 그 앞 구간은 backfill 오염이라 성과로 셀 수 없다.
    assert all(kw.get("start_date") == pr.REPORT_START
               for kw in seen.values())


def test_the_summary_says_when_counting_started(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()])))

    labels = {r[0].value: r[1].value
              for r in load_workbook(path)["요약"].iter_rows()}
    assert labels["집계 시작일"] == pr.REPORT_START


def test_the_summary_reports_the_start_date_actually_used(tmp_path):
    """상수를 찍으면 --start-date 로 바꿔 돌렸을 때 표시가 거짓말을 한다."""
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()]),
                               start_date="2026-08-01"))

    labels = {r[0].value: r[1].value
              for r in load_workbook(path)["요약"].iter_rows()}
    assert labels["집계 시작일"] == "2026-08-01"


# ─── 시장 열 ─────────────────────────────────────────────────
# "어느 시장 상품인가"(NYSE·NASDAQ·ETF)가 없으면 티커만 보고 판단해야 한다.

def test_closed_sheet_carries_the_venue(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(
        _result([_trade(ticker="AAA")], venues={"AAA": "NASDAQ"})))

    assert _cell(path, "주식 청산완료", "시장").value == "NASDAQ"


def test_open_sheet_carries_the_venue(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(
        _result([_held(ticker="SPY")], venues={"SPY": "ETF"})))

    assert _cell(path, "주식 미결포지션", "시장").value == "ETF"


def test_the_venue_is_blank_when_the_archive_did_not_have_it(tmp_path):
    # 2026-08-25 이전 아카이브에는 exchange 열이 없다. 지어내지 않는다.
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade(ticker="AAA")])))

    assert _cell(path, "주식 청산완료", "시장").value in (None, "")


def test_the_venue_sits_next_to_the_ticker(tmp_path):
    # 티커 옆이 아니면 무엇의 시장인지 눈으로 잇기 어렵다.
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(
        _result([_trade(ticker="AAA")], venues={"AAA": "NYSE"})))

    header = [c.value for c in load_workbook(path)["주식 청산완료"][1]]
    assert header[:2] == ["상품티커", "시장"]


# --- quantity comes from the simulation -----------------------------------
# Recomputing it here would let the report disagree with what actually got
# bought, and the same fact would live in two places.

def test_the_row_uses_the_simulated_quantity():
    row = pr.to_row(_trade(qty=7))

    assert row["qty"] == 7
    # 원금도 그 수량으로 계산된다. 100 x 7 = 700, 회수 110 x 7 = 770.
    assert row["gross_usd"] == pytest.approx(70.0)


def test_a_trade_without_a_quantity_is_refused():
    with pytest.raises(ValueError, match="수량"):
        pr.to_row(_trade(qty=None))


def test_the_summary_carries_the_capital_position():
    built = pr.build_rows(_result([_trade()], cash=8_000, capital=10_000))

    assert built["summary"]["capital"] == 10_000
    assert built["summary"]["cash"] == 8_000
    assert built["summary"]["used_pct"] == pytest.approx(20.0)


def test_the_summary_states_the_sizing_rule(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()], cash=8_000, capital=10_000)))

    labels = {r[0].value: r[1].value
              for r in load_workbook(path)["요약"].iter_rows()}

    assert labels["초기 자본($)"] == 10_000
    assert labels["잔여 현금($)"] == 8_000
    assert labels["자본 사용률(%)"] == pytest.approx(20.0)
    assert "1%" in labels["거래당 리스크"]
    assert "20%" in labels["투입 상한"]
    assert "종목당 최대 진입금액($)" not in labels


# --- 현금부족으로 버린 시그널 ------------------------------------------------
# $10,000 에 BUY 가 수십 개면 거절이 규칙이지 예외가 아니다. 요약이 이걸
# 말하지 않으면 매일 보는 산출물에서 버려진 기회가 통째로 안 보인다.

def test_the_summary_counts_the_signals_cash_dropped():
    built = pr.build_rows(_result([_trade()], cash=0, capital=10_000,
                                  skipped_cash=["AR", "INTU", "VLO"]))

    assert built["summary"]["skipped_cash_n"] == 3
    assert built["summary"]["skipped_cash"] == "AR, INTU, VLO"


def test_the_summary_sheet_names_the_dropped_signals(tmp_path):
    path = tmp_path / "r.xlsx"
    _write(path, pr.build_rows(_result([_trade()], cash=0, capital=10_000,
                                       skipped_cash=["AR", "INTU"])))

    labels = {r[0].value: r[1].value
              for r in load_workbook(path)["요약"].iter_rows()}

    assert labels["현금부족 미진입"] == 2
    assert labels["현금부족 미진입 종목"] == "AR, INTU"


def test_the_mail_body_reports_what_cash_dropped():
    built = pr.build_rows(_result([_trade()], cash=0, capital=10_000,
                                  skipped_cash=["AR", "INTU"]))

    body = pr.summary_text({"stocks": built, "etf": None})

    assert "현금부족 2종목" in body


def test_no_cash_shortage_says_nothing():
    """거절이 없으면 줄을 만들지 않는다. 0 을 매일 찍으면 눈이 죽는다."""
    built = pr.build_rows(_result([_trade()], cash=8_000, capital=10_000))

    assert built["summary"]["skipped_cash_n"] == 0
    assert "현금부족" not in pr.summary_text({"stocks": built, "etf": None})

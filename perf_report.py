"""가상매매 성과를 XLSX 리포트로 낸다.

backtest.run() 이 낸 트레이드를 종목당 최대 $1,000 진입으로 환산한다.
R 배수는 리스크 정규화 단위여서 "얼마 벌었나" 에 답하지 못한다.

주식과 ETF 를 각자의 아카이브에서 읽어 시트를 나누고 요약에서만 합계를 낸다.
표본 수가 다른 둘을 한 표에 섞으면 큰 쪽이 작은 쪽의 승률을 가린다.

설계: docs/superpowers/specs/2026-08-19-perf-report-design.md
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import backtest
import console
import exit_rules as er
import history
import mailer
import tracks
import stops
import trade_sim as ts

# 종목당 최대 진입금액(달러). "최대" 인 것은 정수 주로 내림하기 때문이다.
# 진입가가 이 금액보다 비싸면 1주는 산다 - 0주로 두면 고가 종목이 시그널을
# 내도 성과 측정에서 통째로 사라진다.
CAPITAL_USD = 1000

# 성과 집계 시작일. 이 앞의 아카이브(2026-07-31~08-21)는 66% 가 backfill 이라
# 스코어가 미확정 봉 결함에 오염돼 있고, 한국 종목까지 섞여 있다. 깨끗한 live
# 구간부터 다시 센다. 아카이브 자체는 지우지 않는다 - 백테스트는 여전히 전체를
# 볼 수 있어야 한다.
REPORT_START = "2026-08-25"

# 2구획 커스텀 서식. 음수 구획에 - 를 명시하므로 회계 서식의 괄호 표기
# (636.00) 는 나오지 않는다. 정액이 $1,000 이라 센트가 유의미하다.
MONEY_FMT = '#,##0.00;-#,##0.00'
PRICE_FMT = '#,##0.00;-#,##0.00'
QTY_FMT = '#,##0'
RATE_FMT = '#,##0.00'
# 값은 12.34 로 저장하고 서식으로 % 를 붙인다. Excel 기본 0.00% 서식은
# 값이 0.1234 여야 해서, 셀을 직접 읽는 쪽이 100배 틀린다.
PCT_FMT = '0.00"%";-0.00"%"'


def warning_lines(backfill_pct: float) -> list[str]:
    """숫자만 보고 성능으로 오독하는 것을 막는 경고문.

    요약 시트와 메일 본문이 같은 문장을 써야 한다. 한쪽만 고치는 사고를
    막으려고 여기 한 벌만 둔다. 접두사(!!, 들여쓰기)는 각 렌더러가 붙인다 -
    시트는 라벨-값 2열, 본문은 들여쓴 텍스트라 모양이 다르다.
    """
    return [
        "이 리포트는 파이프라인 검증용이다. 시그널 성능의 근거가 아니다.",
        f"아카이브의 {backfill_pct:.0f}% 가 backfill 이라 스코어가 "
        "미확정 봉 결함에 오염돼 있다.",
        "보유 상한 60거래일을 채운 표본이 나오기 전까지 승률·평균은 무의미하다.",
    ]


def to_row(trade, capital: int = CAPITAL_USD,
           costs: ts.Costs = None) -> dict:
    """트레이드 1건을 달러 손익 행으로 환산한다.

    미결 포지션은 청산가 자리에 평가가격(mark_price)이 들어오고 매도비용도
    똑같이 뺀다 - 지금 팔면 손에 남는 돈이 평가액이다.

    두 퍼센트의 분모는 모두 투자원금이다. 순수익%의 분모에 매수비용을
    더하면 두 컬럼을 나란히 비교할 수 없다.
    """
    # 통화가 섞이면 조용히 틀린 금액이 나온다. 아카이브 앞부분에는 원화로
    # 호가되는 한국 종목이 남아 있는데, 그 가격을 달러로 셈하면 GS 의
    # 20,900원 손익이 $20,900 이 된다(2026-08-25 실제로 그렇게 나왔다).
    if trade.market != "US":
        raise ValueError(
            f"{trade.ticker}: 리포트는 USD 전용인데 market={trade.market!r} "
            "이다. 백테스트를 us_only 로 돌려야 한다.")

    costs = costs or ts.Costs()

    # 0주면 손익이 0이라 트레이드가 조용히 사라진다. 1주로 올린다.
    qty = max(1, int(capital // trade.entry_price))
    principal = trade.entry_price * qty

    exit_price = (trade.exit_price if trade.exit_price is not None
                  else trade.mark_price)
    gross = exit_price * qty - principal

    buy_side, sell_side = ts.cost_amount(trade.entry_price, exit_price,
                                         trade.market, costs)
    cost = (buy_side + sell_side) * qty
    net = gross - cost

    return {
        "ticker": trade.ticker,
        "entry_date": trade.entry_date,
        "entry_price": trade.entry_price,
        "exit_date": trade.exit_date,
        "exit_price": exit_price,
        "gross_usd": gross,
        "gross_pct": gross / principal * 100.0,
        "net_usd": net,
        "net_pct": net / principal * 100.0,
        "qty": qty,
        "reason": trade.exit_reason,
        "bars_held": trade.bars_held,
    }


def target_cols(trade) -> dict:
    """미결 포지션의 목표가 관련 네 값. 목표가가 없으면 전부 빈칸이다.

    목표(%) 는 Trade.target_price 에서 되계산한다. 아카이브의 target 정수를
    따로 싣지 않는다 - 같은 사실을 두 곳에 두면 어긋날 수 있고, 목표가가
    유일한 진실이어야 한다.

    달성률은 목표폭 대비 어디까지 왔는지다. 진입가 아래면 음수가 나오고,
    그것이 맞다 - 목표에서 멀어졌다는 뜻이다.
    """
    tp = trade.target_price
    if tp is None:
        return {"target_pct": None, "target_price": None,
                "target_progress_pct": None, "reward_risk": None}

    entry = trade.entry_price
    return {
        "target_pct": (tp / entry - 1) * 100.0,
        "target_price": tp,
        "target_progress_pct": (trade.mark_price - entry) / (tp - entry) * 100.0,
        # 목표폭 ÷ 손절폭. 1 미만이면 목표를 다 채워도 손절 한 번보다 덜 번다.
        "reward_risk": (tp - entry) / trade.r_unit,
    }


def build_rows(result: dict, capital: int = CAPITAL_USD,
               costs: ts.Costs = None, params: er.Params = None,
               start_date: str = REPORT_START) -> dict:
    """청산완료·미결·요약 세 덩어리로 나눈다.

    미결을 승률에 섞지 않는다. trade_sim.summarize 와 같은 원칙이다.

    params 는 미결 포지션의 손절 컬럼에만 쓰인다. 트레일 판정은
    stops.stop_view 에 맡긴다 - 여기서 다시 구현하면 두 벌이 된다.
    """
    costs = costs or ts.Costs()
    params = params or er.Params()
    mark_date = result["newest_bar"]
    # 티커 -> NYSE|NASDAQ|AMEX|ETF. 2026-08-25 이전 아카이브에는 거래소 열이
    # 없어 빈 값이 온다 - 지어내지 않고 그대로 비운다.
    venues = result.get("venues") or {}

    closed, opened = [], []
    for t in result["trades"]:
        if t.is_open:
            row = to_row(t, capital, costs)
            # 미결은 청산일이 없다. 평가 시점을 대신 넣는다.
            row["exit_date"] = mark_date
            sv = stops.stop_view(t, params)
            row["stop"] = sv["stop"]
            # 손절선은 현재가 아래에 있다. 부호가 없으면 상승 여력처럼 읽힌다.
            row["stop_pct"] = -sv["room_pct"]
            row["trail_trigger"] = sv["trail_trigger"]
            row["trail"] = "ON" if sv["trail_active"] else "off"
            row.update(target_cols(t))
            row["venue"] = venues.get(t.ticker, "")
            opened.append(row)
        else:
            row = to_row(t, capital, costs)
            row["venue"] = venues.get(t.ticker, "")
            closed.append(row)

    closed.sort(key=lambda r: (r["exit_date"], r["ticker"]))
    opened.sort(key=lambda r: (r["entry_date"], r["ticker"]))

    wins = sum(1 for r in closed if r["net_usd"] > 0)
    total_rows = result["live_rows"] + result["backfill_rows"]
    return {
        "closed": closed,
        "open": opened,
        "summary": {
            "generated": history.kst_now().strftime("%Y-%m-%d %H:%M KST"),
            "archive_from": result["dates"][0],
            "archive_to": result["dates"][-1],
            "live_rows": result["live_rows"],
            "backfill_rows": result["backfill_rows"],
            "backfill_pct": (result["backfill_rows"] / total_rows * 100.0)
                            if total_rows else 0.0,
            "mark_date": mark_date,
            "failed": result["failed"],
            "closed_n": len(closed),
            "win_rate": (wins / len(closed) * 100.0) if closed else None,
            "gross_usd": sum(r["gross_usd"] for r in closed),
            "net_usd": sum(r["net_usd"] for r in closed),
            # 트레이드별 순수익률의 단순평균이다. 금액 가중이 아니다.
            "avg_net_pct": (sum(r["net_pct"] for r in closed) / len(closed))
                           if closed else None,
            "open_n": len(opened),
            "open_net_usd": sum(r["net_usd"] for r in opened),
            "capital": capital,
            "use_target": params.use_target,
            # 실제로 쓴 시작일을 담는다. 요약이 상수를 찍으면 --start-date 로
            # 바꿔 돌렸을 때 표시가 거짓말을 한다.
            "start_date": start_date,
        },
    }


# (헤더, 행 키, 숫자서식). 앞 9개가 요청받은 컬럼이고, 뒤는 검증용이다 -
# 손익은 가격 x 수량이라 둘이 다 보여야 검산이 된다.
CLOSED_COLS = [
    ("상품티커", "ticker", None),
    ("시장", "venue", None),
    ("진입일자", "entry_date", None),
    ("진입가격", "entry_price", PRICE_FMT),
    ("청산일자", "exit_date", None),
    ("청산가격", "exit_price", PRICE_FMT),
    ("총수익($)", "gross_usd", MONEY_FMT),
    ("총수익(%)", "gross_pct", PCT_FMT),
    ("순수익($)", "net_usd", MONEY_FMT),
    ("순수익(%)", "net_pct", PCT_FMT),
    ("수량", "qty", QTY_FMT),
    ("청산사유", "reason", None),
]

OPEN_COLS = [
    ("상품티커", "ticker", None),
    ("시장", "venue", None),
    ("진입일자", "entry_date", None),
    ("진입가격", "entry_price", PRICE_FMT),
    ("평가기준일", "exit_date", None),
    ("현재가", "exit_price", PRICE_FMT),
    ("평가 총수익($)", "gross_usd", MONEY_FMT),
    ("평가 총수익(%)", "gross_pct", PCT_FMT),
    ("평가 순수익($)", "net_usd", MONEY_FMT),
    ("평가 순수익(%)", "net_pct", PCT_FMT),
    ("수량", "qty", QTY_FMT),
    ("보유봉수", "bars_held", QTY_FMT),
    # 이 전략에는 고정 익절가가 없다. 팔리는 가격은 손절선 하나뿐이고
    # 그것도 고점과 ATR 을 따라 매일 움직이므로 리포트에 실어 둔다.
    ("손절가", "stop", PRICE_FMT),
    ("손절까지(%)", "stop_pct", PCT_FMT),
    ("트레일 발동가", "trail_trigger", PRICE_FMT),
    ("트레일", "trail", None),
    # 손절선이 "어디서 잘리나" 라면 이쪽은 "어디까지 가면 되나" 다.
    # use_target 이 꺼져 있어도 표시한다 - 규칙과 무관한 위치 정보다.
    ("목표(%)", "target_pct", PCT_FMT),
    ("목표가", "target_price", PRICE_FMT),
    ("달성률(%)", "target_progress_pct", PCT_FMT),
    ("위험보상", "reward_risk", RATE_FMT),
]


def _write_sheet(ws, cols, rows) -> None:
    """헤더 + 데이터 행 + 열별 숫자서식. 행이 없어도 헤더는 쓴다."""
    ws.append([title for title, _key, _fmt in cols])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row[key] for _title, key, _fmt in cols])

    for i, (title, _key, fmt) in enumerate(cols, start=1):
        letter = get_column_letter(i)
        if fmt:
            for cell in ws[letter][1:]:
                cell.number_format = fmt
        ws.column_dimensions[letter].width = max(len(title) + 4, 12)


def combined(by_track: dict) -> dict:
    """트랙 요약들을 합계 한 벌로 접는다. 경고문과 총손익이 이걸 쓴다."""
    sums = [(by_track.get(k) or {}).get("summary") for k, _ in TRACK_SHEETS]
    sums = [x for x in sums if x]
    live = sum(x["live_rows"] for x in sums)
    back = sum(x["backfill_rows"] for x in sums)
    closed = [r for k, _ in TRACK_SHEETS
              for r in (by_track.get(k) or EMPTY_TRACK)["closed"]]
    wins = sum(1 for r in closed if r["net_usd"] > 0)
    return {
        "backfill_pct": (back / (live + back) * 100.0) if live + back else 0.0,
        "closed_n": len(closed),
        "win_rate": (wins / len(closed) * 100.0) if closed else None,
        "net_usd": sum(x["net_usd"] for x in sums),
        "open_n": sum(x["open_n"] for x in sums),
        "open_net_usd": sum(x["open_net_usd"] for x in sums),
        "mark_date": max((x["mark_date"] for x in sums if x["mark_date"]),
                         default="-"),
        "failed": sorted({t for x in sums for t in x["failed"]}),
    }


def _track_line(label: str, s: dict) -> str:
    """요약 본문의 트랙 한 줄. 청산이 없으면 건수만 적는다."""
    if not s:
        return f"  {label:5s} 아카이브 없음"
    note = (f"청산 {s['closed_n']}건 승률 {s['win_rate']:.1f}% "
            f"평균 {s['avg_net_pct']:+.2f}%"
            if s["closed_n"] else "청산 0건")
    total = s["net_usd"] + s["open_net_usd"]
    return (f"  {label:5s} ${total:+,.2f}  "
            f"(실현 ${s['net_usd']:+,.2f} · 미실현 ${s['open_net_usd']:+,.2f}"
            f" · {note} · 보유 {s['open_n']}종목)")


def summary_text(by_track: dict) -> str:
    """트랙별 리포트를 콘솔·메일 공용 본문으로 만든다.

    win_rate 는 청산완료가 0건이면 None 이다. 포맷하면 터지므로 건수만 적는다.

    총 수익률(%) 은 싣지 않는다. capital 은 종목당 진입 상한이라 총 투입금
    대비 수익률을 내려면 청산 자금을 재투자하지 않는다는 가정을 몰래
    들여오게 된다.
    """
    c = combined(by_track)
    closed_note = (f"청산 {c['closed_n']}건, 승률 {c['win_rate']:.1f}%"
                   if c["closed_n"] else "청산 0건")
    failed = ", ".join(c["failed"]) if c["failed"] else "없음"
    total = c["net_usd"] + c["open_net_usd"]

    warn = warning_lines(c["backfill_pct"])
    return "\n".join([
        f"!! {warn[0]}",
        f"   {warn[1]}",
        f"   {warn[2]}",
        "",
        f"총 손익      ${total:+,.2f}  ({closed_note})",
        f"  └ 실현     ${c['net_usd']:+,.2f}",
        f"  └ 미실현   ${c['open_net_usd']:+,.2f}  (보유 {c['open_n']}종목)",
        "",
        "[트랙별]",
        *[_track_line(label, (by_track.get(k) or {}).get("summary"))
          for k, label in TRACK_SHEETS],
        "",
        f"평가기준일   {c['mark_date']}",
        f"시세 조회 실패: {failed}",
    ])


def _write_summary(ws, by_track: dict) -> None:
    """라벨-값 2열. 경고를 맨 위에 두고 트랙별 블록 뒤에 합계를 낸다.

    트랙을 한 블록에 합치지 않는 이유는 승률과 평균 때문이다. 표본 수가
    다른 둘을 섞으면 큰 쪽이 작은 쪽을 가린다.
    """
    c = combined(by_track)
    warn = warning_lines(c["backfill_pct"])
    any_summary = next((x for x in
                        ((by_track.get(k) or {}).get("summary")
                         for k, _ in TRACK_SHEETS) if x), None)

    lines = [
        ("!! 경고", warn[0]),
        ("", warn[1]),
        ("", warn[2]),
        ("", ""),
        ("리포트 생성", any_summary["generated"] if any_summary else "-"),
        ("집계 시작일",
         any_summary["start_date"] if any_summary else REPORT_START),
        ("평가기준일", c["mark_date"]),
        ("시세 조회 실패", ", ".join(c["failed"]) if c["failed"] else "없음"),
    ]

    for key, label in TRACK_SHEETS:
        t = (by_track.get(key) or {}).get("summary")
        lines.append(("", ""))
        lines.append((f"[{label}]", ""))
        if not t:
            lines.append(("아카이브", "없음"))
            continue
        lines += [
            ("아카이브 기간", f"{t['archive_from']} ~ {t['archive_to']}"),
            ("live 행수", t["live_rows"]),
            ("backfill 행수", t["backfill_rows"]),
            ("청산 건수", t["closed_n"]),
            ("승률(%)", t["win_rate"]),
            ("누적 총수익($)", t["gross_usd"]),
            ("누적 순수익($)", t["net_usd"]),
            ("평균 순수익률(%)", t["avg_net_pct"]),
            ("보유 건수", t["open_n"]),
            ("평가 순손익($)", t["open_net_usd"]),
        ]

    lines += [
        ("", ""),
        ("[합계]", ""),
        ("청산 건수", c["closed_n"]),
        ("승률(%)", c["win_rate"]),
        ("실현 손익($)", c["net_usd"]),
        ("미실현 손익($)", c["open_net_usd"]),
        ("총 손익($)", c["net_usd"] + c["open_net_usd"]),
        ("", ""),
        ("[가정]", ""),
        ("종목당 최대 진입금액($)", any_summary["capital"] if any_summary else "-"),
        ("매수 수량", "정액 ÷ 진입가, 소수점 내림. 진입가가 더 비싸도 1주는 산다"),
        ("비용", "미국 편도 0.10% · 슬리피지 편도 0.05%"),
        ("통화", "전부 USD. 미국 종목만 보므로 원화 환산을 하지 않는다"),
        ("트랙", "주식과 ETF 는 점수 척도가 달라 시트를 나눈다. 승률·평균을 "
                 "섞으면 표본이 큰 쪽이 작은 쪽을 가린다"),
        ("승률·평균", "청산완료만으로 계산한다. 미결은 제외"),
        ("목표가 익절", "사용함 (목표가 도달 시 청산)"
                      if any_summary and any_summary["use_target"]
                      else "사용 안 함 (--use-target 으로 켬)"),
    ]

    for label, value in lines:
        ws.append([label, value])

    for row in ws.iter_rows(min_col=1, max_col=2):
        label = row[0].value or ""
        if label.endswith("($)"):
            row[1].number_format = MONEY_FMT
        elif label.endswith("(%)"):
            row[1].number_format = PCT_FMT

    ws["A1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80


# (트랙 키, 시트 이름 접두사). 성과도 트랙을 섞지 않는다 - 승률과 평균이
# 무엇의 값인지 알 수 없게 되고, 한쪽 표본이 다른 쪽을 가린다.
TRACK_SHEETS = [("stocks", "주식"), ("etf", "ETF")]

EMPTY_TRACK = {"closed": [], "open": []}


def write_xlsx(path, by_track: dict) -> None:
    """트랙별 청산·미결 시트 4개 + 통합 요약 1개를 쓴다.

    아카이브가 빈 트랙도 시트는 만든다. 시트가 없으면 파일이 깨진 것인지
    트레이드가 없는 것인지 구분되지 않는다.
    """
    wb = Workbook()
    sheets = []
    for key, label in TRACK_SHEETS:
        sheets.append((f"{label} 청산완료", CLOSED_COLS,
                       (by_track.get(key) or EMPTY_TRACK)["closed"]))
    for key, label in TRACK_SHEETS:
        sheets.append((f"{label} 미결포지션", OPEN_COLS,
                       (by_track.get(key) or EMPTY_TRACK)["open"]))

    for i, (title, cols, rows) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        _write_sheet(ws, cols, rows)

    _write_summary(wb.create_sheet("요약"), by_track)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    console.force_utf8()
    p = argparse.ArgumentParser(description="가상매매 성과 누적 리포트")
    p.add_argument("--out-dir", default="reports")
    p.add_argument("--start-date", default=REPORT_START,
                   help=f"집계 시작일 (기본: {REPORT_START})")
    p.add_argument("--capital", type=int, default=CAPITAL_USD,
                   help="종목당 최대 진입금액 USD (기본: 1000)")
    p.add_argument("--mail", action="store_true",
                   help="리포트를 메일로 보낸다 (SMTP_* 환경변수 또는 .env 필요)")
    p.add_argument("--use-target", action="store_true",
                   help="목표가 도달 시 익절한 결과로 리포트를 낸다")
    args = p.parse_args()

    params = er.Params(use_target=args.use_target)

    # 트랙마다 아카이브가 따로다. 한쪽이 비어도 나머지로 리포트를 낸다 -
    # ETF 아카이브는 2026-08-25 분리 시작이라 주식보다 이력이 짧다.
    by_track = {}
    for key, label in TRACK_SHEETS:
        pattern = tracks.history_glob(key)
        # us_only 로 돌린다. 리포트 금액이 전부 달러라 원화로 호가되는
        # 한국 종목이 섞이면 안 된다 - 아카이브 07-31~08-21 구간에 남아 있다.
        result = backtest.run(pattern, params, us_only=True,
                              start_date=args.start_date)
        if not result["dates"]:
            print(f"[!] {label}: 아카이브가 비어 있다 ({pattern})")
            by_track[key] = None
            continue
        by_track[key] = build_rows(result, args.capital, params=params,
                                   start_date=args.start_date)
        print(f"[*] {label}: 트레이드 {len(result['trades'])}건")

    if not any(by_track.values()):
        raise SystemExit("두 트랙 모두 아카이브가 비어 있다")

    stamp = history.kst_now().strftime("%Y-%m-%d")
    path = Path(args.out_dir) / f"perf_{stamp}.xlsx"
    write_xlsx(path, by_track)

    body = summary_text(by_track)
    print(f"{path} 작성 완료")
    print(body)

    if args.mail:
        # 발송 실패를 삼키지 않는다. 조용히 안 보내는 것보다 잡이 실패하는
        # 편이 낫다.
        subject = f"[성과리포트] {stamp} (KST)"
        # 본문은 콘솔에 찍은 것 그대로 + 첨부 안내다.
        mailer.send(subject, f"{body}\n\n상세는 첨부된 {path.name} 참고",
                    [path], **mailer.creds_from_env())
        print(f"메일 발송 완료: {subject}")


if __name__ == "__main__":
    main()
